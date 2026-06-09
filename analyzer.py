# -*- coding: utf-8 -*-
"""PDF 제품평가보고서 분석 로직.
Streamlit app.py에서 import해서 사용합니다.
"""

import os
import re
import tempfile
from io import BytesIO
from datetime import datetime

import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

APP_TITLE = "불량율 분석"
CREATOR_TEXT = "created by 김지연"

OUTPUT_SHEET_RAW = "① 원본백데이터"
OUTPUT_SHEET_DETAIL = "② 불량상세"
OUTPUT_SHEET_SUMMARY = "③ 요약"

RAW_HEADERS = [
    "파일명", "지역1", "지역2", "INS. DATE_시작일", "REPORT NO.", "이전 보고서번호", "바이어_수정", "의뢰업체", "브랜드", "공장",
    "스타일번호", "품명", "검사종류", "ORDER Q'TY", "INSPEC. Q'TY", "PASS Q'TY", "FAIL Q'TY",
    "1차검사수량", "1차합격수량", "1차불합격수량", "2차검사수량", "2차합격수량", "최종불합격수량",
]
for i in range(1, 20):
    RAW_HEADERS += [f"주요불량{i}", f"불량갯수{i}"]

DETAIL_HEADERS = [
    "파일명", "REPORT NO.", "이전 보고서번호", "검사일", "바이어", "의뢰업체", "브랜드", "공장", "지역1", "지역2",
    "스타일번호", "품명", "검사수량(INSPEC)", "원본불량명", "중불량", "경불량", "불량수량"
]


def normalize_space(s):
    # 일부 PDF에서 글자 사이에 들어오는 특수 제어문자/깨진 공백을 정리합니다.
    s = str(s or "")
    s = s.replace("\ufffe", " ").replace("\ufeff", " ").replace("\x01", " ").replace("\x00", " ").replace(" ", " ")
    return re.sub(r"\s+", " ", s).strip()


def compact_label(s):
    # 표 라벨 비교용: 공백/특수문자를 최대한 제거합니다.
    return re.sub(r"[\s\u00a0\ufffe\ufeff\x00-\x1f]+", "", str(s or ""))


def to_int(value):
    if value is None:
        return None
    s = str(value).replace(",", "").replace("PCS", "").replace("pcs", "").strip()
    m = re.search(r"-?\d+", s)
    return int(m.group()) if m else None


def first_match(pattern, text, flags=re.S):
    m = re.search(pattern, text, flags)
    return normalize_space(m.group(1)) if m else None


def extract_ins_start_date(text):
    """
    INS. DATE 값에서 시작일만 추출합니다.
    지원 예: 2024.08.21~22, 2024-12-14~26, 2024.10.11~19,
             2025/2/13~2/25, 2026.03.25~04.04, 20260422
    반환 형식: YYYY.MM.DD
    """
    text = normalize_space(text)
    windows = []
    for m in re.finditer(r"INS\.?\s*DATE", text, re.I):
        windows.append(text[m.end():m.end() + 80])
    windows.append(text)
    patterns = [
        r"([12][0-9]{3})\s*[.\-/]\s*([0-9]{1,2})\s*[.\-/]\s*([0-9]{1,2})",
        r"\b([12][0-9]{3})([01][0-9])([0-3][0-9])\b",
    ]
    for win in windows:
        for pat in patterns:
            m = re.search(pat, win)
            if m:
                y, mth, d = m.groups()
                return f"{int(y):04d}.{int(mth):02d}.{int(d):02d}"
    return None


def text_from_pdf(doc):
    return normalize_space("\n".join(page.get_text("text") for page in doc))


def find_value_in_table(table_rows, label):
    target = compact_label(label)
    for row in table_rows or []:
        row = [normalize_space(c) for c in row]
        compact = [compact_label(c) for c in row]
        for i, c in enumerate(compact):
            if c == target and i + 1 < len(row):
                value = normalize_space(row[i + 1])
                if value:
                    return value
    for r, row in enumerate(table_rows or []):
        compact = [compact_label(c) for c in row]
        for i, c in enumerate(compact):
            if c == target and r + 1 < len(table_rows) and i < len(table_rows[r + 1]):
                value = normalize_space(table_rows[r + 1][i])
                if value:
                    return value
    return None


def find_header_text(table_rows):
    for row in table_rows:
        for c in row:
            c = normalize_space(c)
            if "제품 평가 보고서" in c and "REPORT NO" in c:
                return c
    return ""


def extract_tables(page):
    try:
        return page.find_tables().tables
    except Exception:
        return []


def parse_defects(table_rows, full_text):
    defects = []
    for row in table_rows:
        row = [normalize_space(c) for c in row]
        if not row:
            continue
        name = row[0]
        if name in ("3. 주요 불량 내용", "4. 주요 불량 내용", "주요 불량 내용", "", "합 계", "합계"):
            continue
        if name.startswith("-") or re.match(r"^\d+\.", name):
            continue
        major = to_int(row[1] if len(row) > 1 else None) or 0
        minor = to_int(row[2] if len(row) > 2 else None) or 0
        if major + minor > 0:
            defects.append({"name": name, "major": major, "minor": minor, "qty": major + minor})

    if defects:
        return defects[:19]

    # fallback: OCR/text extraction only. It is less accurate than table extraction.
    defect_names = [
        "바텍 누락", "좌우 비대칭", "제사처리불량", "원단불량", "기름오염",
        "구멍", "이색", "염반", "봉탈", "봉비", "퍼커링", "히까리", "잡사", "오염"
    ]
    for name in defect_names:
        pat = re.compile(re.escape(name) + r"\s+(\d+)(?:\s+(\d+))?(?=\s+(?:\d+(?:,\d+)*|\d+\.))")
        m = pat.search(full_text)
        if m:
            major = to_int(m.group(1)) or 0
            minor = to_int(m.group(2)) or 0
            defects.append({"name": name, "major": major, "minor": minor, "qty": major + minor})
    return defects[:19]



def split_filename_metadata(filename):
    """파일명에서 보조 메타데이터를 추출합니다."""
    base = os.path.splitext(os.path.basename(filename))[0]
    base = re.sub(r"^\(양식변경\)\s*", "", base).strip()
    report = first_match(r"([A-Z][0-9]{3}-[0-9]{2}-[0-9]{5}(?:-[0-9]{2})?)", base, flags=0)
    meta = {"REPORT NO.": report}
    if not report:
        return meta
    tail = base.split(report, 1)[-1].lstrip("_ -")
    parts = [p for p in tail.split("_") if p]
    if len(parts) >= 1:
        meta["의뢰업체"] = parts[0]
    if len(parts) >= 2:
        meta["브랜드"] = parts[1].replace("BEBEDEPINO", "BEBE DE PINO").replace("HENRY_COTTON_S", "HENRY COTTON'S")
    code_candidates = []
    for p in parts[2:]:
        if re.search(r"[A-Za-z]", p) and re.search(r"\d", p) and len(p) >= 6:
            if not re.search(r"^(?:1차|2차|3차|전수|샘플|추가|최초|최조|\d{8})", p):
                code_candidates.append(p)
    if code_candidates:
        meta["스타일번호"] = code_candidates[0]
        try:
            idx = parts.index(code_candidates[0])
            if idx > 2:
                meta["품명"] = " ".join(parts[2:idx])
        except ValueError:
            pass
    m = re.search(r"(20\d{2})(\d{2})(\d{2})", base)
    if m:
        y, mo, d = m.groups()
        meta["INS. DATE_시작일"] = f"{y}.{mo}.{d}"
    return {k: normalize_space(v) for k, v in meta.items() if v}


def apply_filename_fallback(rec, filename):
    meta = split_filename_metadata(filename)
    for k, v in meta.items():
        if rec.get(k) in (None, ""):
            rec[k] = v


def normalize_country_to_region(country):
    country = normalize_space(country).upper()
    return {"VIETNAM": "베트남", "KOREA": "한국", "CHINA": "중국", "MYANMAR": "미얀마"}.get(country, country or None)


def normalize_local_region(region):
    region = normalize_space(region)
    return {"HCMC": "호치민", "DANDONG": "단동"}.get(region, region or None)


def is_bad_table_value(value):
    value = normalize_space(value)
    if not value:
        return True
    bad = ["공장 평가상태", "평가 준비 상태", "사전 완성반평가", "신청업체", "벤더 참관"]
    return any(b in value for b in bad)


def fallback_basic_fields_from_text(rec, full_text, filename):
    text = normalize_space(full_text)
    if not rec.get("REPORT NO."):
        rec["REPORT NO."] = first_match(r"REPORT\s*NO\.?\s*([A-Z0-9\-]+)", text)
    if not rec.get("INS. DATE_시작일"):
        rec["INS. DATE_시작일"] = extract_ins_start_date(text)
    if not rec.get("지역1"):
        rec["지역1"] = normalize_country_to_region(first_match(r"COUNTRY\s+([A-Z]+)", text) or "")
    if not rec.get("스타일번호"):
        meta = split_filename_metadata(filename)
        rec["스타일번호"] = meta.get("스타일번호") or first_match(r"품\s*번\s*([A-Z0-9\-]+)", text)
    m_qty = re.search(r"([0-9,]+)\s*(?:PCS|pcs)\s+([0-9,]+)\s*(?:PCS|pcs)\s+([^\n]*?평가)", full_text, re.I)
    if m_qty:
        if rec.get("ORDER Q'TY") in (None, ""):
            rec["ORDER Q'TY"] = to_int(m_qty.group(1))
        if rec.get("INSPEC. Q'TY") in (None, ""):
            rec["INSPEC. Q'TY"] = to_int(m_qty.group(2))
        if not rec.get("검사종류"):
            rec["검사종류"] = normalize_space(m_qty.group(3))
    apply_filename_fallback(rec, filename)
    return rec

def parse_pdf(pdf_path):
    filename = os.path.basename(pdf_path)
    doc = fitz.open(pdf_path)
    try:
        if len(doc) == 0:
            raise ValueError("빈 PDF")

        full_text = text_from_pdf(doc)
        is_full_inspection = "색상별 제품평가 수량" in full_text or "전수평가" in full_text
        is_sampling = "색상별 오더수량" in full_text or "샘플링 평가" in full_text or "Sampling Plan" in full_text

        if "제품 평가 보고서" not in full_text:
            raise ValueError("문서 제목 확인 실패: '제품 평가 보고서' 텍스트를 찾을 수 없음")
        if not (is_full_inspection or is_sampling):
            raise ValueError("평가 방식 확인 실패: '색상별 제품평가 수량/색상별 오더수량/전수평가/샘플링 평가' 중 식별 가능한 항목이 없음")

        tables = extract_tables(doc[0])
        table0 = tables[0].extract() if len(tables) >= 1 else []
        table1 = tables[1].extract() if len(tables) >= 2 else []
        header_text = find_header_text(table0) or full_text

        rec = {"파일명": filename}
        rec["REPORT NO."] = (first_match(r"REPORT\s*NO\.?\s*([A-Z0-9\-]+)", header_text)
                             or first_match(r"REPORT\s*NO\.?\s*([A-Z0-9\-]+)", full_text)
                             or first_match(r"\(([A-Z0-9\-]+)\)", filename))
        rec["INS. DATE_시작일"] = extract_ins_start_date(header_text) or extract_ins_start_date(full_text)
        country = first_match(r"COUNTRY\s+([A-Z가-힣]+)", header_text) or first_match(r"COUNTRY\s+([A-Z가-힣]+)", full_text)
        rec["지역1"] = normalize_country_to_region(country)

        rec["의뢰업체"] = find_value_in_table(table0, "의뢰업체")
        prev_report_no = find_value_in_table(table0, "이전 보고서번호")
        if prev_report_no and prev_report_no.strip() not in ("-", "–", "—"):
            rec["이전 보고서번호"] = prev_report_no
        else:
            rec["이전 보고서번호"] = ""
        rec["바이어_수정"] = find_value_in_table(table0, "바이어")
        rec["공장"] = find_value_in_table(table0, "공장")
        if is_bad_table_value(rec.get("공장")):
            rec["공장"] = None
        rec["브랜드"] = find_value_in_table(table0, "브랜드")
        region_value = find_value_in_table(table0, "지역")
        # 변경 양식에서는 '지역' 대신 '이전 보고서번호'가 들어갑니다.
        rec["지역2"] = normalize_local_region(region_value) if region_value and not re.match(r"^[A-Z][0-9]{3}-", region_value) else None
        rec["품명"] = find_value_in_table(table0, "품 명")
        rec["스타일번호"] = find_value_in_table(table0, "품 번") or first_match(r"([A-Z]{3,}\d+[A-Z0-9]*)", filename)
        rec["검사종류"] = find_value_in_table(table0, "평가종류")
        rec["ORDER Q'TY"] = to_int(find_value_in_table(table0, "총오더수량"))

        m_qty = re.search(r"([0-9,]+)\s*PCS\s+([0-9,]+)\s*PCS\s+([^\n]*?평가)", full_text)
        second_qty = None
        if m_qty:
            rec["ORDER Q'TY"] = rec.get("ORDER Q'TY") or to_int(m_qty.group(1))
            rec["검사종류"] = rec.get("검사종류") or normalize_space(m_qty.group(3))
            second_qty = to_int(m_qty.group(2))
        if not rec.get("검사종류"):
            rec["검사종류"] = "샘플링 평가" if is_sampling else "전수평가"

        fallback_basic_fields_from_text(rec, full_text, filename)

        if is_full_inspection and not is_sampling:
            color_total = None
            for row in table0:
                row = [normalize_space(c) for c in row]
                if row and row[0] == "합계":
                    color_total = row
                    break
            if not color_total:
                m = re.search(r"합계\s+([0-9,]+)\s+([0-9,]+)\s+(\d+)\s+(\d+)\s+([0-9,]+)\s+([0-9,]+)\s+[0-9.]+%", full_text)
                if m:
                    color_total = ["합계", "", m.group(1), m.group(2), "", "", m.group(3), "", m.group(4), m.group(5), "", m.group(6)]
            if not color_total:
                raise ValueError("색상별 제품평가 수량 합계 행을 찾을 수 없음")

            rec["1차검사수량"] = to_int(color_total[2] if len(color_total) > 2 else None) or second_qty
            rec["1차불합격수량"] = to_int(color_total[3] if len(color_total) > 3 else None)
            rec["최종불합격수량"] = to_int(color_total[6] if len(color_total) > 6 else None)
            rec["2차검사수량"] = rec["1차불합격수량"]
            rec["2차합격수량"] = to_int(color_total[8] if len(color_total) > 8 else None)
            rec["PASS Q'TY"] = to_int(color_total[9] if len(color_total) > 9 else None)
            rec["FAIL Q'TY"] = rec["최종불합격수량"]
            rec["INSPEC. Q'TY"] = to_int(color_total[11] if len(color_total) > 11 else None) or rec["1차검사수량"]
            rec["1차합격수량"] = (rec.get("1차검사수량") or 0) - (rec.get("1차불합격수량") or 0)
        else:
            shipment_total = None
            sample_total = None
            for row in table0:
                row = [normalize_space(c) for c in row]
                if row and row[0] == "합계":
                    nums = [to_int(x) for x in row if to_int(x) is not None]
                    if len(nums) >= 2:
                        shipment_total, sample_total = nums[0], nums[1]
                    break
            if shipment_total is None or sample_total is None:
                m = re.search(r"합계\s+([0-9,]+)\s+([0-9,]+)\s+<\s*합\s*격\s*>", full_text)
                if m:
                    shipment_total = to_int(m.group(1))
                    sample_total = to_int(m.group(2))

            rec["1차검사수량"] = shipment_total or second_qty
            rec["INSPEC. Q'TY"] = sample_total or second_qty
            found = re.search(r"불량발견매수\s+(\d+)\s+(\d+)", full_text)
            rec["1차불합격수량"] = (to_int(found.group(1)) or 0) + (to_int(found.group(2)) or 0) if found else None
            rec["최종불합격수량"] = 0 if ("<합 격>" in full_text or "<합격>" in full_text or "합격허용매수 이내" in full_text) else rec.get("1차불합격수량")
            rec["FAIL Q'TY"] = rec["최종불합격수량"]
            rec["PASS Q'TY"] = (rec.get("INSPEC. Q'TY") or 0) - (rec.get("FAIL Q'TY") or 0) if rec.get("INSPEC. Q'TY") is not None else None
            rec["1차합격수량"] = (rec.get("INSPEC. Q'TY") or 0) - (rec.get("1차불합격수량") or 0) if rec.get("INSPEC. Q'TY") is not None and rec.get("1차불합격수량") is not None else None
            rec["2차검사수량"] = None
            rec["2차합격수량"] = None

        rec["defects"] = parse_defects(table1, full_text)

        required = ["REPORT NO.", "INS. DATE_시작일", "스타일번호", "INSPEC. Q'TY"]
        missing = [k for k in required if rec.get(k) in (None, "")]
        if missing:
            snapshot_keys = ["REPORT NO.", "INS. DATE_시작일", "의뢰업체", "브랜드", "공장", "스타일번호", "품명", "검사종류", "ORDER Q'TY", "INSPEC. Q'TY"]
            snapshot = "; ".join(f"{k}={rec.get(k)!r}" for k in snapshot_keys)
            raise ValueError("필수 필드 누락: " + ", ".join(missing) + " | extracted: " + snapshot)
        return rec
    finally:
        doc.close()


def make_workbook(records, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = OUTPUT_SHEET_RAW
    detail = wb.create_sheet(OUTPUT_SHEET_DETAIL)
    summary = wb.create_sheet(OUTPUT_SHEET_SUMMARY)

    title_fill = PatternFill("solid", fgColor="E2F0D9")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(["▶ 원본 백데이터 PDF 자동 추출"])
    ws.append(RAW_HEADERS)
    for rec in records:
        base_col_count = 23  # RAW_HEADERS에서 주요불량 컬럼이 시작되기 전까지의 컬럼 수
        row = [rec.get(h) for h in RAW_HEADERS[:base_col_count]]
        defect_values = []
        for d in rec.get("defects", []):
            defect_values += [d["name"], d["qty"]]
        while len(defect_values) < 38:
            defect_values += [None, None]
        ws.append(row + defect_values[:38])

    detail.append(["▶ 불량항목 상세"])
    detail.append(DETAIL_HEADERS)
    for rec in records:
        for d in rec.get("defects", []):
            detail.append([
                rec.get("파일명"), rec.get("REPORT NO."), rec.get("이전 보고서번호"), rec.get("INS. DATE_시작일"), rec.get("바이어_수정"), rec.get("의뢰업체"),
                rec.get("브랜드"), rec.get("공장"), rec.get("지역1"), rec.get("지역2"), rec.get("스타일번호"), rec.get("품명"),
                rec.get("INSPEC. Q'TY"), d["name"], d["major"], d["minor"], d["qty"]
            ])

    summary.append(["▶ PDF 통합 요약"])
    summary.append(["생성일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    summary.append(["총 보고서 수", len(records)])
    summary.append(["총 검사 수량", sum((r.get("INSPEC. Q'TY") or 0) for r in records)])
    summary.append(["총 1차불량수량", sum((r.get("1차불합격수량") or 0) for r in records)])
    summary.append(["총 최종불합격수량", sum((r.get("최종불합격수량") or 0) for r in records)])

    for sheet in [ws, detail, summary]:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = border
                if cell.row == 1:
                    cell.font = Font(bold=True, size=12)
                    cell.fill = title_fill
                elif cell.row == 2:
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
        for col in range(1, sheet.max_column + 1):
            max_len = 0
            for cell in sheet[get_column_letter(col)]:
                max_len = max(max_len, len(str(cell.value or "")))
            sheet.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 10), 32)
        sheet.freeze_panes = "A3"

    wb.save(output_path)





def workbook_to_bytes(records):
    """분석 결과 records를 엑셀 바이너리(bytes)로 반환합니다."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    try:
        make_workbook(records, tmp_path)
        with open(tmp_path, "rb") as f:
            return f.read()
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass


def analyze_uploaded_files(uploaded_files):
    """Streamlit UploadedFile 목록을 분석합니다.

    Returns
    -------
    dict
        {
          "records": list[dict],
          "failed_files": list[tuple[str, str]],
          "logs": list[str],
          "excel_bytes": bytes | None,
        }
    """
    pdf_files = [f for f in uploaded_files if getattr(f, "name", "").lower().endswith(".pdf")]
    total_pdf_count = len(pdf_files)
    records = []
    failed_files = []
    logs = []

    if total_pdf_count == 0:
        return {"records": [], "failed_files": [], "logs": ["No PDF files found."], "excel_bytes": None}

    with tempfile.TemporaryDirectory() as tmpdir:
        for idx, uploaded in enumerate(pdf_files, start=1):
            filename = os.path.basename(uploaded.name)
            pdf_path = os.path.join(tmpdir, filename)
            with open(pdf_path, "wb") as f:
                f.write(uploaded.getbuffer())

            try:
                rec = parse_pdf(pdf_path)
                records.append(rec)
                logs.append(f"({idx}/{total_pdf_count}) loaded: {filename}")
            except Exception as e:
                reason = str(e)
                failed_files.append((filename, reason))
                logs.append(f"({idx}/{total_pdf_count}) load failed: {filename}")
                logs.append(f"  reason: {reason}")

    logs.append("")
    logs.append(f"{total_pdf_count} files exists.")
    logs.append(f"{len(records)} files are successfully loaded.")
    logs.append(f"{len(failed_files)} files are not loaded correctly.")

    if failed_files:
        logs.append("")
        logs.append("not loaded files:")
        for filename, reason in failed_files:
            logs.append(f"{filename} | reason: {reason}")

    excel_bytes = workbook_to_bytes(records) if records else None
    return {"records": records, "failed_files": failed_files, "logs": logs, "excel_bytes": excel_bytes}
